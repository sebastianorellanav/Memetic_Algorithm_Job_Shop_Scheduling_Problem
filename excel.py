import pandas as pd
import openpyxl as op

def create_excel(name_sheet):
    columns_excel = pd.DataFrame({'Instancia': ['tamano poblacion', 'iteraciones', 'porcentaje de preservacion', 'umbral convergencia', 'iteraciones busqueda local', 'Makespan', 'Tiempo']})
    columns_excel = columns_excel.transpose()
    # crear el objeto ExcelWriter
    excelBook = op.load_workbook('resultados.xlsx')
    with pd.ExcelWriter('resultados.xlsx') as writer:
        # Save your file workbook as base
        writer.book = excelBook
        writer.sheets = dict((ws.title, ws) for ws in excelBook.worksheets)

        # Now here add your new sheets
        columns_excel.to_excel(writer,name_sheet, startrow = writer.sheets[name_sheet].max_row, header=False)

        # Save the file
        writer.save()

def write_result(result, name_sheet):
    result = result.transpose()
    excelBook = op.load_workbook('resultados.xlsx')
    with pd.ExcelWriter('resultados.xlsx') as writer:
        # Save your file workbook as base
        writer.book = excelBook
        writer.sheets = dict((ws.title, ws) for ws in excelBook.worksheets)

        # Now here add your new sheets
        result.to_excel(writer,name_sheet, startrow = writer.sheets[name_sheet].max_row, header=False)

        # Save the file
        writer.save()
